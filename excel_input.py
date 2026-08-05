from pathlib import Path
from typing import Any

from openpyxl import load_workbook

#different things to call the same column header in the Excel file. More logic late for extra stuff
HEADER_ALIASES = {
    "mpn": {
        "mpn",
        "manufacturerpartnumber",
        "manufacturerpn",
        "mfrpartnumber",
        "partnumber",
    },
    "supplier": {
        "supplier",
        "suppliername",
        "distributor",
        "distributorname",
    },
}

#Convert an Excel cell into clean text.
def clean_cell(value: Any):
    if value is None:
        return ""

    # Excel occasionally returns whole numbers as floats.
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    return str(value).strip().strip("'\"")

# Normalize a heading by removing spaces and punctuation.
def normalize_header(value: Any):
    header = clean_cell(value).casefold()

    return "".join(
        character
        for character in header
        if character.isalnum()
    )

# Find one MPN column and all supplier columns.
def find_column_positions(headers):
    mpn_column: int | None = None
    supplier_columns: list[int] = []

    for column_index, header in enumerate(headers):
        normalized_header = normalize_header(header)

        if normalized_header in HEADER_ALIASES["mpn"]:
            mpn_column = column_index

        elif (
            normalized_header.startswith("supplier")
            or normalized_header.startswith("distributor")
        ):
            supplier_columns.append(column_index)

    if mpn_column is None:
        raise ValueError("Could not find an MPN column.")

    if not supplier_columns:
        raise ValueError(
            "Could not find any supplier columns."
        )

    return mpn_column, supplier_columns

#Read MPN and supplier data from a user-provided Excel workbook.
def read_bom(file_path: str | Path,sheet_name: str | None = None,):
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Excel file does not exist: {path}"
        )

    if not path.is_file():
        raise ValueError(
            f"The supplied path is not a file: {path}"
        )

    if path.suffix.casefold() != ".xlsx":
        raise ValueError(
            "Only .xlsx Excel files are currently supported."
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name is None:
            sheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(
                    workbook.sheetnames
                )

                raise ValueError(
                    f"Worksheet '{sheet_name}' was not found. "
                    f"Available sheets: {available_sheets}"
                )

            sheet = workbook[sheet_name]

        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)

        if headers is None:
            raise ValueError(
                "The selected worksheet is empty."
            )

        mpn_column, supplier_columns = find_column_positions(
            headers
        )

        bom_rows: list[dict[str, Any]] = []

        for excel_row_number, row in enumerate(
            rows,
            start=2,
        ):
            mpn_value = (
                row[mpn_column]
                if mpn_column < len(row)
                else None
            )

            mpn = clean_cell(mpn_value)

            suppliers: list[str] = []
            seen_suppliers: set[str] = set()

            for supplier_column in supplier_columns:
                supplier_value = (
                    row[supplier_column]
                    if supplier_column < len(row)
                    else None
                )

                supplier = clean_cell(supplier_value)

                if not supplier:
                    continue

                normalized_supplier = supplier.casefold()

                if normalized_supplier in seen_suppliers:
                    continue

                seen_suppliers.add(normalized_supplier)
                suppliers.append(supplier)

            # Ignore completely empty rows.
            if not mpn and not suppliers:
                continue

            if not mpn:
                raise ValueError(
                    f"Excel row {excel_row_number} "
                    "is missing an MPN."
                )

            if not suppliers:
                raise ValueError(
                    f"Excel row {excel_row_number} "
                    "has no suppliers."
                )

            bom_rows.append(
                {
                    "excel_row": excel_row_number,
                    "mpn": mpn,
                    "suppliers": suppliers,
                }
            )

        if not bom_rows:
            raise ValueError(
                "The worksheet contains no BOM data rows."
            )

        return bom_rows

    finally:
        workbook.close()

# Build variables for a Nexar supMultiMatch query.
# Duplicate MPNs are queried only once.
def build_nexar_variables(bom_rows,result_limit: int = 1):
    if result_limit < 1:
        raise ValueError(
            "result_limit must be at least 1."
        )

    queries: list[dict[str, Any]] = []
    seen_mpns: set[str] = set()

    for bom_row in bom_rows:
        mpn = bom_row["mpn"]
        normalized_mpn = mpn.casefold()

        if normalized_mpn in seen_mpns:
            continue

        seen_mpns.add(normalized_mpn)

        queries.append(
            {
                "mpn": mpn,
                "start": 0,
                "limit": result_limit,
            }
        )

    return {
        "queries": queries,
    }