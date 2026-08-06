# holds functions for reading an Excel file and extracting MPN and supplier numbers and names,
# used to build variables for a Nexar supMultiMatch query (request sent to API).
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

# possible header names for the columns we are searching for
# more logic later on for symbols ect
HEADER_ALIASES = {
    "mpn": {
        "mpn",
        "manufacturerpartnumber",
        "manufacturerpn",
        "mfrpartnumber",
        "partnumber",
    },
    "supplier": {
        "supplier",
        "suppliername",
        "distributor",
        "distributorname",
    },
}

# Convert an Excel cell into clean text for use.
def clean_cell(value):
    if value is None:
        return ""

    # Excel occasionally returns whole numbers as floats.
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    return str(value).strip().strip("'\"")

# Normalize a heading by removing spaces and punctuation.
def normalize_header(value):
    header = clean_cell(value).casefold()

    normalized_header = ""

    for character in header:
        if character.isalnum():
            normalized_header += character

    return normalized_header

# Find one MPN column and all supplier columns, return their indices.
# headers parameter is the first row of the Excel sheet which you get from the read_bom function
def find_column_positions(headers):
    #set to None in case no MPN or supplier columns found
    mpn_column: int | None = None
    supplier_columns: list[int] | None = None

    for column_index, header in enumerate(headers):
        normalized_header = normalize_header(header)

        if normalized_header in HEADER_ALIASES["mpn"]:
            mpn_column = column_index
            continue

        for supplier_header in HEADER_ALIASES["supplier"]:
            supplier_number = normalized_header.removeprefix(
                supplier_header
            )

            if (normalized_header == supplier_header or supplier_number.isdigit()):
                if supplier_columns is None:
                    supplier_columns = []
                supplier_columns.append(column_index)
                break

    if mpn_column is None:
        raise ValueError("Could not find an MPN column.")
    if supplier_columns is None:
        raise ValueError("Could not find any supplier columns.")

    return mpn_column, supplier_columns

# Read MPN and supplier data from Excel sheet
# sheet name is optional, if not provided the active sheet will be used (change to auto look at all sheets)
def read_bom(file_path,sheet_name: str | None = None,):
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Excel file does not exist: {path}"
        )

    if path.suffix.casefold() != ".xlsx":
        raise ValueError(
            f"The supplied path is not an Excel .xlsx file: {path}"
        )

    try:
        #excel file opened in read-only 
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )

        if sheet_name is None:
            sheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(
                    workbook.sheetnames
                )

                raise ValueError(
                    f"Worksheet '{sheet_name}' was not found. "
                    f"Available sheets: {available_sheets}"
                )

            sheet = workbook[sheet_name]

        # Iterator for excel rows
        rows = sheet.iter_rows(values_only=True)
        # gets the first row of the excel sheet and updates iterator
        headers = next(rows, None)

        if headers is None:
            raise ValueError(
                "The selected worksheet is empty."
            )

        mpn_column, supplier_columns = find_column_positions(headers)

        # dict has keys for excel row number, mpn, and suppliers (might add column to mpn and supplier changing pair to datastructure)
        bom_rows: list[dict[str, Any]] = []

        # rows start at 1 but its 2 because the first row was headers
        for excel_row_number, row in enumerate(rows,start=2):
            if mpn_column < len(row):
                mpn = row[mpn_column]
            else:
                mpn = None

            mpn = clean_cell(mpn)

            # Keep suppliers in one list because each row may have any number of suppliers.
            # Separate keys like supplier_1, supplier_2, etc. would make iteration harder later.
            suppliers: list[str] = []
            # in case of duplicate suppliers in the same row
            seen_suppliers: set[str] = set()

            for supplier_column in supplier_columns:
                if supplier_column < len(row):
                    supplier = row[supplier_column]
                else:
                    supplier = None

                supplier = clean_cell(supplier)

                if not supplier:
                    continue

                # Normalize only for comparison. Keep the original supplier spelling
                normalized_supplier = "".join(
                    character
                    for character in supplier.casefold()
                    if character.isalnum()
                )

                if normalized_supplier in seen_suppliers:
                    continue

                seen_suppliers.add(normalized_supplier)
                suppliers.append(supplier)

            # Ignore completely empty rows.
            if not mpn and not suppliers:
                continue

            if not mpn:
                raise ValueError(
                    f"Excel row {excel_row_number} "
                    "is missing an MPN."
                )

            if not suppliers:
                raise ValueError(
                    f"Excel row {excel_row_number} "
                    "has no suppliers."
                )

            bom_rows.append(
                {
                    "excel_row": excel_row_number,
                    "mpn": mpn,
                    "suppliers": suppliers,
                }
            )

        if not bom_rows:
            raise ValueError(
                "The worksheet contains no relevant BOM data rows."
            )

        return bom_rows

    finally:
        workbook.close()

# Build variables for a Nexar supMultiMatch query (avoids duplicate queries).
def build_nexar_variables(bom_rows,result_limit: int = 1):
    if result_limit < 1:
        raise ValueError(
            "result_limit must be at least 1."
        )

    queries: list[dict[str, Any]] = []
    seen_mpns: set[str] = set()

    for bom_row in bom_rows:
        mpn = bom_row["mpn"]
        normalized_mpn = mpn.casefold()

        if normalized_mpn in seen_mpns:
            continue

        seen_mpns.add(normalized_mpn)

        queries.append(
            {
                "mpn": mpn,
                "start": 0,
                "limit": result_limit,
            }
        )

    return {
        "queries": queries,
    }