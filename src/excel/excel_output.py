from pathlib import Path
from typing import Any

#for excel control
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import src.excel.excel_input as excel_input
import src.nexar.utilities.sort as sort

#Convert supplier stock results into Excel text.
def format_stock(supplier_results):
    if not supplier_results:
        return "No data returned"

    stock_values: list[str] = []
    multiple_offers = len(supplier_results) > 1

    for offer_number, supplier_result in enumerate(
        supplier_results,
        start=1,
    ):
        stock = supplier_result.get("stock")

        if stock is None:
            stock_text = "Not provided"
        else:
            stock_text = str(stock)

        if multiple_offers:
            stock_text = (
                f"Offer {offer_number}: {stock_text}"
            )

        stock_values.append(stock_text)

    return "\n".join(stock_values)


#Convert supplier pricing results into Excel text.
def format_pricing(supplier_results):
    if not supplier_results:
        return "No data returned"

    pricing_values: list[str] = []
    multiple_offers = len(supplier_results) > 1

    for offer_number, supplier_result in enumerate(
        supplier_results,
        start=1,
    ):
        prices = supplier_result.get("pricing") or []
        price_breaks: list[str] = []

        for price in prices:
            quantity = price.get("quantity")
            unit_price = price.get("unit_price")
            currency = price.get("currency") or ""

            if quantity is None:
                quantity_text = "Quantity not provided"
            else:
                quantity_text = f"{quantity}+"

            if unit_price is None:
                price_text = "Price not provided"
            else:
                price_text = f"{unit_price} {currency}".strip()

            price_breaks.append(
                f"{quantity_text}: {price_text}"
            )

        if price_breaks:
            pricing_text = "; ".join(price_breaks)
        else:
            pricing_text = "No pricing returned"

        if multiple_offers:
            pricing_text = (
                f"Offer {offer_number}: {pricing_text}"
            )

        pricing_values.append(pricing_text)

    return "\n".join(pricing_values)

#Match each Excel MPN with its corresponding Nexar result group.
def build_match_groups_by_mpn(nexar_data,nexar_variables):
    queries = nexar_variables.get("queries") or []
    match_groups = nexar_data.get("supMultiMatch") or []

    match_groups_by_mpn: dict[str, dict[str, Any]] = {}

    for query_index, query in enumerate(queries):
        mpn = excel_input.clean_cell(
            query.get("mpn")
        )

        if query_index < len(match_groups):
            match_group = match_groups[query_index]
        else:
            match_group = {}

        match_groups_by_mpn[mpn.casefold()] = match_group

    return match_groups_by_mpn

#Create a copy of the BOM with supplier stock and pricing columns.
def write_bom_results(input_file,output_file,nexar_data,nexar_variables,sheet_name):
    input_path = Path(input_file)
    output_path = Path(output_file)

    workbook = load_workbook(input_path)

    try:
        if sheet_name is None:
            sheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet '{sheet_name}' was not found."
                )

            sheet = workbook[sheet_name]

        original_max_column = sheet.max_column

        headers = tuple(
            sheet.cell(
                row=1,
                column=column_number,
            ).value
            for column_number in range(
                1,
                original_max_column + 1,
            )
        )

        mpn_column, supplier_columns = (
            excel_input.find_column_positions(headers)
        )

        match_groups_by_mpn = build_match_groups_by_mpn(
            nexar_data=nexar_data,
            nexar_variables=nexar_variables,
        )

        output_columns: list[dict[str, int]] = []

        for supplier_number, supplier_column in enumerate(
            supplier_columns,
            start=1,
        ):
            supplier_header = excel_input.clean_cell(
                headers[supplier_column]
            )

            if not supplier_header:
                supplier_header = (
                    f"Supplier {supplier_number}"
                )

            stock_column = (original_max_column+ ((supplier_number - 1) * 2)+ 1)

            pricing_column = stock_column + 1

            sheet.cell(row=1,column=stock_column,).value = f"{supplier_header} Stock"

            sheet.cell(row=1,column=pricing_column,).value = f"{supplier_header} Pricing"

            sheet.cell(
                row=1,
                column=stock_column,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

            sheet.cell(
                row=1,
                column=pricing_column,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

            sheet.column_dimensions[
                get_column_letter(stock_column)
            ].width = 20

            sheet.column_dimensions[
                get_column_letter(pricing_column)
            ].width = 55

            output_columns.append(
                {
                    "supplier_column": (
                        supplier_column + 1
                    ),
                    "stock_column": stock_column,
                    "pricing_column": pricing_column,
                }
            )

        for excel_row_number in range(
            2,
            sheet.max_row + 1,
        ):
            mpn = excel_input.clean_cell(
                sheet.cell(
                    row=excel_row_number,
                    column=mpn_column + 1,
                ).value
            )

            if not mpn:
                continue

            match_group = match_groups_by_mpn.get(
                mpn.casefold()
            )

            if match_group:
                part_data = {
                    "supMultiMatch": [match_group]
                }
            else:
                part_data = {
                    "supMultiMatch": []
                }

            for output_column in output_columns:
                supplier_name = excel_input.clean_cell(
                    sheet.cell(
                        row=excel_row_number,
                        column=output_column[
                            "supplier_column"
                        ],
                    ).value
                )

                if not supplier_name:
                    continue

                supplier_results = (
                    sort.get_supplier_stock_and_pricing(
                        nexar_data=part_data,
                        supplier_name=supplier_name,
                    )
                )

                stock_cell = sheet.cell(
                    row=excel_row_number,
                    column=output_column[
                        "stock_column"
                    ],
                )

                pricing_cell = sheet.cell(
                    row=excel_row_number,
                    column=output_column[
                        "pricing_column"
                    ],
                )

                stock_cell.value = format_stock(
                    supplier_results
                )

                pricing_cell.value = format_pricing(
                    supplier_results
                )

                stock_cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

                pricing_cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(output_path)

        return output_path

    finally:
        workbook.close()