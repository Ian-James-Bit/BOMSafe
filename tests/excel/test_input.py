# TODO: add more asserts once more test excel sheets are added to relevant test
import src.excel.input as input
import pytest
from pathlib import Path
from openpyxl import load_workbook

@pytest.fixture
def path(): 
    return "uploads/final_bomsafe_test.xlsx"
 
@pytest.fixture
def get_sheet(path): 
    workbook = load_workbook(
        filename=Path(path),
        read_only=True,
        data_only=True,
    )
    yield workbook.active

    workbook.close()

@pytest.fixture
def get_header_row(get_sheet):
    rows = get_sheet.iter_rows(values_only=True)
    return next(rows, None)

def test_clean_cell():
    assert input.clean_cell("'MPN'") == "MPN"
    assert input.clean_cell("  1726480102  ") == "1726480102"
    assert input.clean_cell(99.0) == "99"
    assert input.clean_cell("Digi-Key") == "Digi-Key"
    assert input.clean_cell(None) == ""

def test_normalize_header():
    assert input.normalize_header("Manufacturer Part Number") == "manufacturerpartnumber"
    assert input.normalize_header("Supplier 1") == "supplier1"
    assert input.normalize_header("MFR-Part_Number") == "mfrpartnumber"

# TODO: add more asserts once more test excel sheets are added
def test_find_column_positions(get_header_row):
    assert input.find_column_positions(get_header_row) == (0, [1, 2])

# TODO: add more asserts once more test excel sheets are added
def test_read_bom(path):
    assert input.read_bom(path) == [
        {
            "excel_row": 2,
            "mpn": "1726480102",
            "suppliers": ["Mouser", "DigiKey"],
        }
    ]

# TODO: add more asserts once more test excel sheets are added
def test_build_nexar_variables(path):
    assert input.build_nexar_variables(input.read_bom(path)) == {
    "queries": [
            {
                "mpn": "1726480102",
                "start": 0,
                "limit": 1,
            }
        ]
    }